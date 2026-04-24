import json
import os
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


CATEGORY_ORDER = [
    "Technical Skills",
    "Domain/Tools/Process",
    "Team Management",
    "People Management Skills",
    "Communication Skills",
    "Behavioral Skills",
]

BAND_HEADERS = [
    "Band - T\n(0-1 year)",
    "Band - A1\n(1-2 years)",
    "Band - A2\n(2-4 years)",
    "Band - A3\n(4-6 years)",
    "Band - B1\n(6-7 years)",
    "Band - B2\n(7-9 years)",
    "Band - B3\n(9-11 years)",
]


def sanitize_sheet_name(name: str) -> str:
    bad = ['\\', '/', '*', '[', ']', ':', '?']
    cleaned = name
    for ch in bad:
        cleaned = cleaned.replace(ch, " ")
    cleaned = " ".join(cleaned.split()).strip()
    return cleaned[:31] if cleaned else "Role"


def group_skills_by_category(skills: List[Dict]) -> Dict[str, List[str]]:
    grouped = {cat: [] for cat in CATEGORY_ORDER}
    for s in skills or []:
        cat = (s.get("category") or "").strip()
        name = (s.get("name") or "").strip()
        if not name:
            continue
        if cat not in grouped:
            grouped.setdefault(cat, [])
        grouped[cat].append(name)

    # Deduplicate while preserving order
    for cat, items in grouped.items():
        seen = set()
        deduped = []
        for item in items:
            k = item.lower()
            if k not in seen:
                seen.add(k)
                deduped.append(item)
        grouped[cat] = deduped
    return grouped


def style_cell(cell, fill=None, bold=False, align_center=False):
    if fill:
        cell.fill = fill
    if bold:
        cell.font = Font(bold=True)
    if align_center:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    else:
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def apply_borders(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = border


def build_role_sheet(wb: Workbook, role_name: str, skills: List[Dict]):
    ws = wb.create_sheet(title=sanitize_sheet_name(role_name))
    grouped = group_skills_by_category(skills)

    # Columns A:J
    col_widths = {
        1: 22,  # Skill Area
        2: 28,  # Skills Concepts
        3: 12, 4: 12, 5: 12, 6: 12, 7: 12, 8: 12, 9: 12, 10: 12
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    yellow = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")
    peach = PatternFill(fill_type="solid", start_color="F4B183", end_color="F4B183")
    green = PatternFill(fill_type="solid", start_color="92D050", end_color="92D050")

    # Row heights
    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 35

    # Top header row
    ws["A1"] = "Role"
    style_cell(ws["A1"], fill=yellow, bold=True, align_center=True)

    ws.merge_cells("B1:C1")
    ws["B1"] = role_name
    style_cell(ws["B1"], fill=yellow, bold=True, align_center=True)

    ws.merge_cells("D1:J1")
    ws["D1"] = (
        "0. Not required\n"
        "1. Basic: Knows the concepts\n"
        "2. Intermediate: Knows the application areas\n"
        "3. Advanced: Can use them\n"
        "4. Master: Used them extensively"
    )
    style_cell(ws["D1"], fill=yellow, bold=True, align_center=True)

    # Proficiency ask row
    ws.merge_cells("D2:J2")
    ws["D2"] = "Proficiency Band Ask\n[0, 1, 2, 3, 4]"
    style_cell(ws["D2"], fill=peach, bold=True, align_center=True)

    # Column header row
    ws["A3"] = "Skill Area"
    ws["B3"] = "Skills Concepts"
    style_cell(ws["A3"], fill=green, bold=True, align_center=True)
    style_cell(ws["B3"], fill=green, bold=True, align_center=True)

    for i, band in enumerate(BAND_HEADERS, start=3):
        cell = ws.cell(row=3, column=i, value=band)
        style_cell(cell, fill=green, bold=True, align_center=True)

    current_row = 4

    for category in CATEGORY_ORDER:
        items = grouped.get(category, [])

        # Category row
        ws.cell(row=current_row, column=1, value=category)
        style_cell(ws.cell(row=current_row, column=1), bold=True)
        current_row += 1

        # Skills rows
        if items:
            for skill_name in items:
                ws.cell(row=current_row, column=2, value=skill_name)
                current_row += 1
        else:
            # Keep one blank line if no skills in category
            current_row += 1

        # Spacer line between categories
        current_row += 1

    end_row = current_row - 1
    apply_borders(ws, 1, end_row, 1, 10)
    ws.freeze_panes = "A4"


def convert_json_to_formatted_excel(input_json: str, output_excel: str) -> None:
    input_path = Path(input_json)
    output_path = Path(output_excel)

    with input_path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    for item in data:
        role = (item.get("role") or "Role").strip()
        skills = item.get("skills", []) or []
        build_role_sheet(wb, role, skills)

    wb.save(output_path)
    print(f"Saved formatted workbook: {output_path}")


if __name__ == "__main__":
    script_dir = Path(__file__).resolve().parent
    project_root = script_dir.parent
    jd_root_dir = Path(os.getenv("JD_ROOT_DIR", str(project_root))).expanduser().resolve()

    default_input = jd_root_dir / "extracted_skills_results.json"
    default_output = jd_root_dir / "extracted_skills_results_formatted.xlsx"

    convert_json_to_formatted_excel(
        os.getenv("SKILL_JSON_PATH", str(default_input)),
        os.getenv("SKILL_EXCEL_PATH", str(default_output)),
    )